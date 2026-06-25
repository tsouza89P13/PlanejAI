from io import BytesIO
from unicodedata import normalize

import openpyxl


def _normalizar_texto(valor) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip().lower()
    texto = normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return " ".join(texto.split())


def _normalizar_valor(valor):
    return str(valor).strip() if valor is not None else None


def _canonico(valor, mapa: dict[str, str]):
    chave = _normalizar_texto(valor)
    return mapa.get(chave, _normalizar_valor(valor))


CRITICIDADE_CANONICA = {"alta": "Alta", "media": "Média", "baixa": "Baixa"}
STATUS_CANONICO = {"ativo": "Ativo", "inativo": "Inativo"}
TIPO_CANONICO = {"inspecao": "Inspeção", "preventiva": "Preventiva", "preditiva": "Preditiva"}
DISCIPLINA_CANONICA = {"mecanica": "Mecânica", "eletrica": "Elétrica", "lubrificacao": "Lubrificação"}
FREQUENCIA_CANONICA = {
    "diaria": "Diária",
    "semanal": "Semanal",
    "quinzenal": "Quinzenal",
    "mensal": "Mensal",
    "bimestral": "Bimestral",
    "trimestral": "Trimestral",
    "quadrimestral": "Quadrimestral",
    "semestral": "Semestral",
    "anual": "Anual",
}


def _estilizar_template(ws, colunas):
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col_idx, (_, label, largura) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = largura


def _adicionar_instrucoes(wb, instrucoes: list[str], largura: int = 70) -> None:
    ws_info = wb.create_sheet("Instruções")
    for row_idx, texto in enumerate(instrucoes, 1):
        ws_info.cell(row=row_idx, column=1, value=texto)
    ws_info.column_dimensions["A"].width = largura


def _adicionar_exemplo(ws, exemplo: list) -> None:
    from openpyxl.styles import PatternFill

    exemplo_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    for col_idx, valor in enumerate(exemplo, 1):
        cell = ws.cell(row=2, column=col_idx, value=valor)
        cell.fill = exemplo_fill


def _workbook_bytes(wb) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def gerar_template_equipamentos() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipamentos"
    colunas = [
        ("codigo", "Código *", 15),
        ("descricao", "Descrição *", 40),
        ("local", "Local", 30),
        ("area", "Área", 25),
        ("criticidade", "Criticidade *", 15),
        ("status", "Status *", 12),
    ]
    _estilizar_template(ws, colunas)
    _adicionar_instrucoes(
        wb,
        [
            "INSTRUÇÕES DE PREENCHIMENTO",
            "",
            "* Campos obrigatórios",
            "",
            "CRITICIDADE: Alta | Média | Baixa",
            "STATUS: Ativo | Inativo",
            "",
            "OBSERVAÇÕES:",
            "- Não altere os nomes das colunas",
            "- Você pode usar fórmulas nas células",
            "- Linhas vazias são ignoradas",
            "- O código do equipamento deve ser único",
            "- Se o código já existir no banco, o registro será ATUALIZADO",
        ],
        largura=60,
    )
    _adicionar_exemplo(ws, ["EQ-001", "Bomba Centrífuga", "Sala de Bombas", "Utilidades", "Alta", "Ativo"])
    return _workbook_bytes(wb)


def gerar_template_planos() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planos"
    colunas = [
        ("codigo_equipamento", "Código Equipamento *", 22),
        ("nome", "Nome do Plano *", 35),
        ("descricao", "Descrição", 40),
        ("tipo_intervencao", "Tipo Intervenção *", 20),
        ("disciplina", "Disciplina *", 15),
        ("frequencia", "Frequência *", 15),
        ("duracao_hh", "Duração (HH) *", 15),
        ("prioridade", "Prioridade *", 12),
        ("grupo_parada", "Grupo de Parada", 20),
        ("janela_inicio", "Janela Início *", 15),
        ("janela_fim", "Janela Fim *", 12),
        ("status", "Status *", 12),
    ]
    _estilizar_template(ws, colunas)
    _adicionar_instrucoes(
        wb,
        [
            "INSTRUÇÕES DE PREENCHIMENTO",
            "",
            "* Campos obrigatórios",
            "",
            "TIPO INTERVENÇÃO: Inspeção | Preventiva | Preditiva",
            "DISCIPLINA: Mecânica | Elétrica | Lubrificação",
            "FREQUÊNCIA: Diária | Semanal | Quinzenal | Mensal | Bimestral | Trimestral | Quadrimestral | Semestral | Anual",
            "PRIORIDADE: 1 (máxima) a 5 (mínima)",
            "JANELA INÍCIO: número da semana (1 a 52)",
            "JANELA FIM: número da semana (1 a 52)",
            "STATUS: Ativo | Inativo",
            "",
            "OBSERVAÇÕES:",
            "- Código Equipamento deve existir no banco ou ser importado antes",
            "- Não altere os nomes das colunas",
            "- Você pode usar fórmulas nas células",
            "- Linhas vazias são ignoradas",
            "- Se o plano já existir (mesmo nome + mesmo equipamento), será ATUALIZADO",
            "- Grupo de Parada é opcional; deixe em branco se não aplicável",
        ],
    )
    _adicionar_exemplo(
        ws,
        [
            "EQ-001",
            "Inspeção Visual",
            "Inspeção geral",
            "Inspeção",
            "Mecânica",
            "Mensal",
            2.0,
            2,
            "Parada Mensal",
            1,
            52,
            "Ativo",
        ],
    )
    return _workbook_bytes(wb)


def _cabecalho(ws, mapa_colunas: dict[str, str]) -> dict[str, int]:
    cabecalho = {}
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            campo = mapa_colunas.get(_normalizar_texto(cell.value))
            if campo:
                cabecalho[campo] = cell.column - 1
    return cabecalho


def _get(row, cabecalho: dict[str, int], campo: str):
    idx = cabecalho.get(campo)
    if idx is None:
        return None
    return _normalizar_valor(row[idx])


def _get_num(row, cabecalho: dict[str, int], campo: str, tipo=float):
    idx = cabecalho.get(campo)
    if idx is None:
        return None
    val = row[idx]
    try:
        return tipo(val) if val is not None else None
    except (ValueError, TypeError):
        return None


def importar_equipamentos(arquivo_bytes: bytes, session) -> dict:
    from models import Equipamento

    try:
        wb = openpyxl.load_workbook(BytesIO(arquivo_bytes), data_only=True)
        ws = wb.active
    except Exception as e:
        return {"criados": 0, "atualizados": 0, "erros": [f"Erro ao abrir arquivo: {str(e)}"]}

    mapa_colunas = {
        "codigo *": "codigo",
        "codigo": "codigo",
        "descricao *": "descricao",
        "descricao": "descricao",
        "local": "local",
        "area": "area",
        "criticidade *": "criticidade",
        "criticidade": "criticidade",
        "status *": "status",
        "status": "status",
    }
    cabecalho = _cabecalho(ws, mapa_colunas)
    obrigatorios = ["codigo", "descricao", "criticidade", "status"]
    faltantes = [campo for campo in obrigatorios if campo not in cabecalho]
    if faltantes:
        return {"criados": 0, "atualizados": 0, "erros": [f"Colunas obrigatórias ausentes: {', '.join(faltantes)}. Use o template oficial."]}

    criados = atualizados = 0
    erros = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell for cell in row if cell is not None):
            continue
        codigo = _get(row, cabecalho, "codigo")
        descricao = _get(row, cabecalho, "descricao")
        criticidade = _canonico(_get(row, cabecalho, "criticidade"), CRITICIDADE_CANONICA) or "Média"
        status = _canonico(_get(row, cabecalho, "status"), STATUS_CANONICO) or "Ativo"
        local = _get(row, cabecalho, "local")
        area = _get(row, cabecalho, "area")

        linha_erros = []
        if not codigo:
            linha_erros.append("Código obrigatório")
        if not descricao:
            linha_erros.append("Descrição obrigatória")
        if criticidade not in CRITICIDADE_CANONICA.values():
            linha_erros.append(f"Criticidade inválida: '{criticidade}'")
        if status not in STATUS_CANONICO.values():
            linha_erros.append(f"Status inválido: '{status}'")
        if linha_erros:
            erros.append(f"Linha {row_idx}: {'; '.join(linha_erros)}")
            continue

        existente = session.query(Equipamento).filter(Equipamento.codigo == codigo).first()
        if existente:
            existente.descricao = descricao
            existente.local = local
            existente.area = area
            existente.criticidade = criticidade
            existente.status = status
            atualizados += 1
        else:
            session.add(Equipamento(codigo=codigo, descricao=descricao, local=local, area=area, criticidade=criticidade, status=status))
            criados += 1

    try:
        session.commit()
    except Exception as e:
        session.rollback()
        return {"criados": 0, "atualizados": 0, "erros": [f"Erro ao salvar no banco: {str(e)}"]}
    return {"criados": criados, "atualizados": atualizados, "erros": erros}


def importar_planos(arquivo_bytes: bytes, session) -> dict:
    from models import Equipamento, Plano

    try:
        wb = openpyxl.load_workbook(BytesIO(arquivo_bytes), data_only=True)
        ws = wb.active
    except Exception as e:
        return {"criados": 0, "atualizados": 0, "erros": [f"Erro ao abrir arquivo: {str(e)}"]}

    mapa_colunas = {
        "codigo equipamento *": "codigo_equipamento",
        "codigo equipamento": "codigo_equipamento",
        "nome do plano *": "nome",
        "nome do plano": "nome",
        "descricao": "descricao",
        "tipo intervencao *": "tipo_intervencao",
        "tipo intervencao": "tipo_intervencao",
        "disciplina *": "disciplina",
        "disciplina": "disciplina",
        "frequencia *": "frequencia",
        "frequencia": "frequencia",
        "duracao (hh) *": "duracao_hh",
        "duracao (hh)": "duracao_hh",
        "prioridade *": "prioridade",
        "prioridade": "prioridade",
        "grupo de parada": "grupo_parada",
        "janela inicio *": "janela_inicio",
        "janela inicio": "janela_inicio",
        "janela fim *": "janela_fim",
        "janela fim": "janela_fim",
        "status *": "status",
        "status": "status",
    }
    cabecalho = _cabecalho(ws, mapa_colunas)
    obrigatorios = ["codigo_equipamento", "nome", "tipo_intervencao", "disciplina", "frequencia", "duracao_hh", "prioridade", "janela_inicio", "janela_fim", "status"]
    faltantes = [campo for campo in obrigatorios if campo not in cabecalho]
    if faltantes:
        return {"criados": 0, "atualizados": 0, "erros": [f"Colunas obrigatórias ausentes: {', '.join(faltantes)}. Use o template oficial."]}

    equipamentos_cache = {eq.codigo: eq for eq in session.query(Equipamento).all()}
    criados = atualizados = 0
    erros = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell for cell in row if cell is not None):
            continue

        codigo_eq = _get(row, cabecalho, "codigo_equipamento")
        nome = _get(row, cabecalho, "nome")
        descricao = _get(row, cabecalho, "descricao")
        tipo = _canonico(_get(row, cabecalho, "tipo_intervencao"), TIPO_CANONICO)
        disciplina = _canonico(_get(row, cabecalho, "disciplina"), DISCIPLINA_CANONICA)
        frequencia = _canonico(_get(row, cabecalho, "frequencia"), FREQUENCIA_CANONICA)
        duracao_hh = _get_num(row, cabecalho, "duracao_hh", float)
        prioridade = _get_num(row, cabecalho, "prioridade", int)
        grupo_parada = _get(row, cabecalho, "grupo_parada")
        janela_inicio = _get_num(row, cabecalho, "janela_inicio", int)
        janela_fim = _get_num(row, cabecalho, "janela_fim", int)
        status = _canonico(_get(row, cabecalho, "status"), STATUS_CANONICO) or "Ativo"

        linha_erros = []
        if not codigo_eq or codigo_eq not in equipamentos_cache:
            linha_erros.append(f"Equipamento '{codigo_eq}' não encontrado no banco")
        if not nome:
            linha_erros.append("Nome do plano obrigatório")
        if tipo not in TIPO_CANONICO.values():
            linha_erros.append(f"Tipo inválido: '{tipo}'")
        if disciplina not in DISCIPLINA_CANONICA.values():
            linha_erros.append(f"Disciplina inválida: '{disciplina}'")
        if frequencia not in FREQUENCIA_CANONICA.values():
            linha_erros.append(f"Frequência inválida: '{frequencia}'")
        if duracao_hh is None or duracao_hh <= 0:
            linha_erros.append("Duração (HH) deve ser maior que zero")
        if prioridade is None or prioridade < 1 or prioridade > 5:
            linha_erros.append("Prioridade deve estar entre 1 e 5")
        if janela_inicio is None or janela_fim is None or not (1 <= janela_inicio <= 52) or not (1 <= janela_fim <= 52):
            linha_erros.append("Janelas devem estar entre 1 e 52")
        elif janela_inicio > janela_fim:
            linha_erros.append("Janela início não pode ser maior que janela fim")
        if status not in STATUS_CANONICO.values():
            linha_erros.append(f"Status inválido: '{status}'")
        if linha_erros:
            erros.append(f"Linha {row_idx}: {'; '.join(linha_erros)}")
            continue

        equipamento = equipamentos_cache[codigo_eq]
        existente = session.query(Plano).filter(Plano.equipamento_id == equipamento.id, Plano.nome == nome).first()
        plano = existente or Plano()
        plano.equipamento_id = equipamento.id
        plano.nome = nome
        plano.descricao = descricao
        plano.tipo_intervencao = tipo
        plano.disciplina = disciplina
        plano.frequencia = frequencia
        plano.duracao_hh = duracao_hh
        plano.prioridade = prioridade
        plano.grupo_parada = grupo_parada if grupo_parada else None
        plano.janela_inicio = janela_inicio
        plano.janela_fim = janela_fim
        plano.status = status
        session.add(plano)
        if existente:
            atualizados += 1
        else:
            criados += 1

    try:
        session.commit()
    except Exception as e:
        session.rollback()
        return {"criados": 0, "atualizados": 0, "erros": [f"Erro ao salvar no banco: {str(e)}"]}
    return {"criados": criados, "atualizados": atualizados, "erros": erros}
